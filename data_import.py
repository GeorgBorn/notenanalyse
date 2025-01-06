"""Sammlung aller Funktionen, die für den Datenimport aus den Exceltabellen notwendig sind."""

import openpyxl
import re

subjects = {"D":"Deutsch","De":"Deutsch","Deu":"Deutsch","Deutsch":"Deutsch",
            "M":"Mathematik","Ma":"Mathematik","Mat":"Mathematik","Math":"Mathematik","Mathe":"Mathematik","Mathematik":"Mathematik",
            "E":"Englisch","En":"Englisch","Eng":"Englisch","Engl":"Englisch","Englisch":"Englisch",
            "Ek":"Erdkunde","Erd":"Erdkunde","Erdk":"Erdkunde","Erdkunde":"Erdkunde",
            "G":"Geschichte","Ge":"Geschichte","Geschichte":"Geschichte",
            "Bio":"Biologie","Bi":"Biologie","Biol":"Biologie","Biologie":"Biologie",
            "Ch":"Chemie","Che":"Chemie","Chemie":"Chemie",
            "Ph":"Physik","Phy":"Physik","Phys":"Physik","Physik":"Physik",
            "Phi":"Philosophie","Phil":"Philosophie","Philo":"Philosophie","Philosophie":"Philosophie",
            "WiPo":"Wirtschaft-Politik","WIPO":"Wirtschaft-Politik","WirtPol":"Wirtschaft-Politik",
            "Sp":"Sport","Spo":"Sport","Sport":"Sport",
            "WPU":"Wahlpflichtunterricht","WahlPflicht":"Wahlpflichtunterricht"}

def schuelernamensliste(dokumentenpfad):
    """Gibt eine Liste mit 2er-Tupeln mit Name, Vorname und Zeilennummer aus"""
    wb = openpyxl.load_workbook(dokumentenpfad)
    ws = wb['Tabelle1']
    letzte_zeile = ws.max_row
    schueler_namen = []
    for zeile in range(3, letzte_zeile+1):
        if ws.cell(row=zeile, column=2).value != None and not(re.search(r"[A-Z][A-Z]", str(ws.cell(row=zeile, column=2).value))) and not(re.search(r"[A-Z][1-9]", str(ws.cell(row=zeile, column=2).value))):
            schueler_namen.append((ws.cell(row=zeile, column=2).value,ws.cell(row=zeile, column=3).value,zeile))
    return schueler_namen

def faecherliste(dokumentenpfad):
    """Gibt eine Liste mit 4er-Tupeln mit Fach, Spaltennummer aus"""
    wb = openpyxl.load_workbook(dokumentenpfad)
    ws = wb['Tabelle1']
    letzte_spalte = ws.max_column
    faecher = []
    for spalte in range(5,letzte_spalte+1,2):
        if ws.cell(row=1, column=spalte).value in subjects.keys():
            faecher.append((subjects[ws.cell(row=1, column=spalte).value],spalte))
    return faecher

def jahr_halbjahr_klasse(dokumentenpfad):
    """Gibt die Klasse, das Jahr und das Halbjahr der Notenliste in einem Tupel aus."""
    wb = openpyxl.load_workbook(dokumentenpfad)
    ws = wb['Tabelle1']
    letzte_spalte = ws.max_column
    klasse = ws.cell(row=2, column=3).value
    for spalte in range(4,letzte_spalte):
        if re.search(r"20[1-9][1-9]", str(ws.cell(row=2, column=spalte).value)):
            date = ws.cell(row=2, column=spalte).value
            date_year = date.strftime("%Y")
    if ws.cell(row=4, column=4).value == "2.hbj" or "2.hj":
        if ws.cell(row=4, column=5).value:
            halbjahr = "2.HJ"
        else:
            halbjahr = "1.HJ"
    return (klasse, halbjahr, date_year)

